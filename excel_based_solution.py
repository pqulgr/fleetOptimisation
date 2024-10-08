import streamlit as st
import pandas as pd
from export_analyzer import ExportAnalyzer
from fleet_estimator import FleetEstimator

import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment


OPTION_FREQ = {
    "Jours": "D",
    "Semaines": "W",
    "Mois": "MS",
    "Trimestres": "QS"
}

def export_results_to_excel(analyzer, fleet_estimator):
    output = io.BytesIO()
    workbook = Workbook()
    
    # Feuille pour les données brutes et les prédictions
    ws_data = workbook.active
    ws_data.title = "Données et Prédictions"
    
    # Combiner les données brutes et les prédictions
    df_real = analyzer.df.rename(columns={'y': 'y_real'})
    df_pred = analyzer.forecast.rename(columns={'yhat1': 'y_pred'})
    combined_data = pd.merge(df_real, df_pred[['ds', 'y_pred']], on='ds', how='outer')
    combined_data = combined_data.sort_values('ds')
    
    for r in dataframe_to_rows(combined_data, index=False, header=True):
        ws_data.append(r)
    
    # Ajout d'un graphique pour les données réelles et les prédictions
    chart = LineChart()
    chart.title = "Données réelles vs Prédictions"
    chart.y_axis.title = "Nombre de colis"
    chart.x_axis.title = "Date"
    
    # Données réelles
    y_col_real = ws_data['B']  # Colonne 'y_real' (données réelles)
    data_real = Reference(ws_data, min_col=2, min_row=1, max_row=len(y_col_real))
    chart.add_data(data_real, titles_from_data=True)
    
    # Prédictions
    y_col_pred = ws_data['C']  # Colonne 'y_pred' (prédictions)
    data_pred = Reference(ws_data, min_col=3, min_row=1, max_row=len(y_col_pred))
    chart.add_data(data_pred, titles_from_data=True)
    
    dates = Reference(ws_data, min_col=1, min_row=2, max_row=len(combined_data)+1)
    chart.set_categories(dates)
    
    ws_data.add_chart(chart, "E2")
    
    # Feuille pour les métriques du modèle
    ws_metrics = workbook.create_sheet("Métriques du modèle")
    metrics = analyzer.calculate_model_metrics()
    ws_metrics.append(["Métrique", "Valeur"])
    for key, value in metrics.items():
        ws_metrics.append([key, value])
    
    # Mise en forme du tableau des métriques
    for cell in ws_metrics["A"] + ws_metrics[1]:
        cell.font = Font(bold=True)
    
    # Feuille pour l'estimation de la flotte
    ws_fleet = workbook.create_sheet("Estimation de la flotte")
    summary_table = fleet_estimator.create_summary_table()
    for r in dataframe_to_rows(summary_table, index=False, header=True):
        ws_fleet.append(r)
    
    ws_fleet.append([])
    ws_fleet.append(["Meilleur délai de retour", fleet_estimator.best_reverse_time])
    ws_fleet.append(["Nombre d'emballages estimé", fleet_estimator.fleet_size])
    ws_fleet.append(["Coût estimé", f"{fleet_estimator.results[fleet_estimator.best_reverse_time]['cost']:.0f}€"])
    ws_fleet.append(["Maximum de sacs stocké en entrepôt", f"{fleet_estimator.results[fleet_estimator.best_reverse_time]['max_in_stockage']:.0f} Sacs"])
    
    # Mise en forme du tableau de la flotte
    for cell in ws_fleet["A"] + ws_fleet[1]:
        cell.font = Font(bold=True)
    
    # Feuille pour l'analyse des composantes
    ws_components = workbook.create_sheet("Analyse des composantes")
    components = analyzer.forecast[['ds', 'trend', 'season_yearly', 'season_weekly']]
    for r in dataframe_to_rows(components, index=False, header=True):
        ws_components.append(r)
    
    # Ajout d'un graphique pour les composantes
    chart = LineChart()
    chart.title = "Décomposition des composantes de la prédiction"
    chart.y_axis.title = "Valeur"
    chart.x_axis.title = "Date"
    
    for col in range(2, 5):
        data = Reference(ws_components, min_col=col, min_row=1, max_row=len(components)+1)
        chart.add_data(data, titles_from_data=True)
    
    dates = Reference(ws_components, min_col=1, min_row=2, max_row=len(components)+1)
    chart.set_categories(dates)
    
    ws_components.add_chart(chart, "F2")
    
    # Ajustement de la largeur des colonnes
    for ws in workbook.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde du fichier Excel
    workbook.save(output)
    output.seek(0)
    
    return output
def initialize_session_state():
    if 'app_state' not in st.session_state:
        st.session_state.app_state = {
            'analyzer': ExportAnalyzer(),
            'data_loaded': False,
            'model_trained': False,
            'fleet_estimated': False,
            'fleet_estimator': None,
            'fleet_size': None,
            'cost_option': None,
            'cost_params': None
        }

def get_cost_options():
    cost_params = {}
    cost_option = st.selectbox("Sélectionnez une option de coût", ("Coût basé sur les points de collecte", "Coût basé sur le poids"))
    cost_option = "Option 1" if cost_option == "Coût basé sur les points de collecte" else "Option 2"
    
    if cost_option == "Option 1":
        st.write("Délai avant reverse (jours)")
        col1, col2 = st.columns(2)
        with col1:
            min_reverse_time = st.number_input("Min", min_value=1, max_value=400, value=7, step=1)
        with col2:
            max_reverse_time = st.number_input("Max", min_value=1, max_value=400, value=20, step=1)
        if min_reverse_time > max_reverse_time:
            min_reverse_time, max_reverse_time = max_reverse_time, min_reverse_time
        cost_params["reverse_time"] = (min_reverse_time, max_reverse_time)
        cost_params["nb_locations"] = st.number_input("Nombre de destinations", min_value=1, step=1, value=120)
        cost_params["cost_emballage"] = st.number_input("Coût d'achat des emballages", min_value=0.0, step=0.01, value=3.5)
        cost_params["cost_location"] = st.number_input("Coût de récupération à un point relai", min_value=0.0, step=0.1, value=5.0)
        cost_params["cost_per_demand"] = st.number_input("Coût par envoi", min_value=0.0, step=0.1, value=0.12)
    elif cost_option == "Option 2":
        cost_params["reverse_time"] = st.number_input("Délai avant reverse (jours)", min_value=1, step=1, value=3)
        cost_params["cost_emballage"] = st.number_input("Coût d'achat des emballages", min_value=0.0, step=0.01, value=2.0)
        cost_params["poids_sac"] = st.number_input("Poids moyen", min_value=0.0, step=0.1, value=1.0)
        cost_params["cost_per_demand"] = st.number_input("Coût par envoi par Kg", min_value=0.0, step=0.1, value=1.0)
        cost_params["cost_per_return"] = st.number_input("Coût par retour par Kg", min_value=0.0, step=0.1, value=1.0)

    return cost_option, cost_params

def display_data_analysis():
    st.subheader("Aperçu des données")
    st.table(st.session_state.app_state['analyzer'].df.head(8))
    st.write(f"Nombre total d'enregistrements: {len(st.session_state.app_state['analyzer'].df)}")
    st.write(f"Période couverte: du {st.session_state.app_state['analyzer'].df['ds'].min().date()} au {st.session_state.app_state['analyzer'].df['ds'].max().date()}")
    
    st.plotly_chart(st.session_state.app_state['analyzer'].plot_input_data())
    
    col1, col2 = st.columns(2)
    with col1:
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_average_by_weekday())
    with col2:
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_monthly_evolution())

def display_model_results():
    try:
        st.session_state.app_state['analyzer'].display_model_summary()
        st.session_state.app_state['analyzer'].display_model_metrics()
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_forecast())
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_holiday_impact())
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_seasonal_trends())
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_component_importance())
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_forecast_components())
        st.plotly_chart(st.session_state.app_state['analyzer'].plot_residuals())
        
    except Exception as e:
        st.error(f"Erreur lors de la création des graphiques : {str(e)}")

def estimate_fleet_size():
    with st.spinner("Estimation de la taille de la flotte en cours..."):
        true_data = st.session_state.app_state['analyzer'].df
        predictions = st.session_state.app_state['analyzer'].forecast
        
        st.session_state.app_state['fleet_estimator'] = FleetEstimator(
            true_data, 
            predictions, 
            st.session_state.app_state['cost_option'], 
            st.session_state.app_state['cost_params']
        )
        st.session_state.app_state['fleet_size'] = st.session_state.app_state['fleet_estimator'].estimate_fleet_size()
        st.session_state.app_state['fleet_estimated'] = True

def display_fleet_estimation():
    st.session_state.app_state['fleet_estimator'].display_results()

def main_excel():
    st.title("Analyse de données et prédiction des exportations")
    
    initialize_session_state()
    
    uploaded_file = st.file_uploader("Choisissez un fichier Excel ou CSV", type=["xlsx", "csv"])
    
    if uploaded_file is not None:
        file_type = uploaded_file.name.split('.')[-1].lower()
        
        date_column = st.text_input("Nom de la colonne de date", "Date")
        colis_column = st.text_input("Nom de la colonne de nombre de colis", "nombre de colis")
        
        sheet_name = None
        if file_type == 'xlsx':
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            selected_sheet = st.selectbox("Choisissez une feuille (optionnel)", ["Première feuille"] + sheet_names)
            sheet_name = None if selected_sheet == "Première feuille" else selected_sheet
        
        if st.button("Charger les données"):
            if st.session_state.app_state['analyzer'].load_data(uploaded_file, date_column, colis_column, sheet_name):
                st.session_state.app_state['data_loaded'] = True
                st.success("Données chargées et préparées avec succès!")
            else:
                st.error("Erreur lors du chargement des données.")
        
        if st.session_state.app_state['data_loaded']:
            display_data_analysis()
            
            st.subheader("Configuration du modèle NeuralProphet")
            st.session_state.app_state['analyzer'].model_components['trend'] = st.checkbox("Inclure la tendance", value=True)
            st.session_state.app_state['analyzer'].model_components['artificial_noise'] = st.checkbox("Inclure du bruit", value=True)

            #Frequence de la prédiction
            col1, col2 = st.columns(2)
            with col1:
                selected_freq = st.selectbox("Choisissez la période de prédiction", list(OPTION_FREQ.keys()))
                st.session_state.app_state['analyzer'].model_components['freq'] = OPTION_FREQ[selected_freq]
            with col2:
                st.write(f"Les prédictions se font par {selected_freq.lower()}")

            # Ajuster le nombre de périodes futures en fonction de la fréquence sélectionnée
            if selected_freq == "Jours":
                future_periods_max = 1000
            elif selected_freq == "Semaines":
                future_periods_max = 200
            elif selected_freq == "Mois":
                future_periods_max = 60
            elif selected_freq == "Trimestres":
                future_periods_max = 20
            else:  # Années
                future_periods_max = 10

            st.session_state.app_state['analyzer'].model_params['future_periods'] = st.slider(
                f"Nombre de {selected_freq.lower()} futurs à prédire", 
                min_value=10, 
                max_value=future_periods_max, 
                value=min(st.session_state.app_state['analyzer'].model_params['future_periods'], future_periods_max)
            )

            st.session_state.app_state['analyzer'].model_params['epochs'] = st.slider("Nombre d'époques", 
                                                        min_value=10, max_value=300, 
                                                        value=st.session_state.app_state['analyzer'].model_params['epochs'])
            
            st.session_state.app_state['analyzer'].use_country_holidays = st.session_state.app_state['analyzer'].get_event_features()

            if st.button("Entraîner le modèle et faire des prédictions"):
                with st.spinner("Entraînement du modèle en cours..."):
                    st.session_state.app_state['analyzer'].train_model()
                st.session_state.app_state['model_trained'] = True
                st.success("Modèle entraîné et prédictions générées avec succès!")
            
        if st.session_state.app_state['model_trained']:
            display_model_results()

            st.subheader("Estimation de la taille de la flotte")
            st.session_state.app_state['cost_option'], st.session_state.app_state['cost_params'] = get_cost_options()

            if st.button("Estimer la taille de la flotte"):
                estimate_fleet_size()

        if st.session_state.app_state['fleet_estimated']:
            display_fleet_estimation()
            #if st.button("Exporter les résultats en Excel"):
            excel_file = export_results_to_excel(st.session_state.app_state['analyzer'], st.session_state.app_state['fleet_estimator'])
            st.download_button(
                label="Télécharger le fichier Excel",
                data=excel_file,
                file_name="resultats_analyse_prediction.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        initialize_session_state()

if __name__ == "__main__":
    main_excel()